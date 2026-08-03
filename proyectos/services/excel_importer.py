import logging
from datetime import datetime
from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from .asignador_aulas import AsignadorAulas

from proyectos.models import (
    ModalidadAcademica,
    PeriodoAcademico,
    MomentoAcademico,
    Sede,
    Edificio,
    Aula,
    Docente,
    Clase,
    ImportacionProgramacion,
)

from .mappers import ExcelMapper
from .programacion_backup import capturar_programacion
from .validators import ExcelValidator

logger = logging.getLogger(__name__)


class ExcelImporter:

    MAPA_SEDES = {
        "BUC": "CU Bucaramanga",
        "S": "CU Bucaramanga",
        "CUC": "CU Cúcuta",
        "OCA": "CU Ocaña",
    }

    MAPA_EDIFICIOS = {
        "DJCB": "Diego Jaramillo",
        "SJEB": "San Juan Eudes",
        "FARLAB": "Biblioteca",
        "COLROS": "Colegio del Rosario",
    }

    MAPA_AULAS = {

        "SAL BIBLIO": "FarLab",

        "FARLAB": "FarLab",

        "MAKERLAB": "FarLab",

        "200": "201",

    }

    DIAS = {
        "lunes": 1,
        "martes": 2,
        "miercoles": 3,
        "jueves": 4,
        "viernes": 5,
        "sabado": 6,
        "domingo": 7,
    }

    EDIFICIOS_VIRTUALES = {"VIRTU", "VIRTUAL", "SINCR"}

    ETIQUETAS_OMITIDAS = {
        "referencia": "Referencias sin programación (Consultar ...)",
        "inactiva": "Clases marcadas como inactivas",
        "sin_programacion": "Registros sin horario asignado",
        "sin_dia": "Registros sin día de clase",
        "sin_horas": "Registros sin horario completo",
        "sin_ubicacion": "Registros sin espacio definido",
    }

    def __init__(self, archivo, institucion, usuario=None):

        self.archivo = archivo
        self.institucion = institucion
        self.usuario = usuario

        self.errores = []
        self.total = 0
        self.filas_revisadas = 0
        self.filas_validas = 0
        self.filas_invalidas = 0
        self.filas_omitidas = 0
        self.omitidas_por_motivo = {
            motivo: 0 for motivo in self.ETIQUETAS_OMITIDAS
        }
        self.muestras_omitidas = []
        self.sedes_creadas = []
        self.edificios_creados = []
        self.aulas_creadas = []
        self.docentes_creados = []

    def _resultado(self, total=None):
        return {
            "total": self.total if total is None else total,
            "errores": self.errores,
            "filas_revisadas": self.filas_revisadas,
            "filas_validas": self.filas_validas,
            "filas_invalidas": self.filas_invalidas,
            "filas_omitidas": self.filas_omitidas,
            "resumen_omitidas": [
                {"etiqueta": self.ETIQUETAS_OMITIDAS[motivo], "total": total}
                for motivo, total in self.omitidas_por_motivo.items()
                if total
            ],
            "muestras_omitidas": self.muestras_omitidas,
            "sedes_creadas": self.sedes_creadas,
            "edificios_creados": self.edificios_creados,
            "aulas_creadas": self.aulas_creadas,
            "docentes_creados": self.docentes_creados,
        }

    def previsualizar(self):
        """Valida un archivo sin conservar ningún cambio en la base de datos."""
        with transaction.atomic():
            resultado = self.importar()
            transaction.set_rollback(True)
        return resultado

    @staticmethod
    def _tiene_valor(valor):
        return valor is not None and str(valor).strip() != ""

    @classmethod
    def _dia_marcado(cls, valor):
        return cls._tiene_valor(valor) and str(valor).strip().upper() not in {
            "0", "N", "NO", "FALSE",
        }

    @classmethod
    def _texto(cls, valor):
        return str(valor or "").strip()

    def _omitir_fila(self, numero_fila, nrc, motivo):
        """Registra filas informativas o aún no programadas, sin tratarlas como error."""
        self.filas_omitidas += 1
        self.omitidas_por_motivo[motivo] += 1
        if len(self.muestras_omitidas) < 8:
            identificador = nrc or "sin identificador"
            self.muestras_omitidas.append(
                f"Fila {numero_fila} ({identificador}): "
                f"{self.ETIQUETAS_OMITIDAS[motivo].lower()}."
            )

    def convertir_hora(self, valor):

        if valor is None:
            return None

        valor = str(valor).replace(".0", "").strip()

        while len(valor) < 4:
            valor = "0" + valor

        return datetime.strptime(
            valor,
            "%H%M"
        ).time()

    @transaction.atomic
    def importar(self):

        logger.info("SIGEA IMPORTADOR: Inicio de importación")
        logger.info(f"Institución: {self.institucion.nombre} (ID: {self.institucion.id})")

        workbook = load_workbook(
            self.archivo,
            data_only=True
        )

        hoja = workbook.active
        logger.info(f"Hoja activa: {hoja.title}")
        logger.info(f"Total filas en hoja: {hoja.max_row}")

        cabeceras = [

            str(c.value).strip().upper()

            if c.value else ""

            for c in hoja[1]

        ]

        logger.info(f"Cabeceras detectadas: {cabeceras}")

        indices, faltantes = ExcelMapper.obtener_indices(
            cabeceras
        )

        logger.info(f"Índices mapeados: {indices}")
        logger.info(f"Columnas faltantes: {faltantes}")

        errores = ExcelValidator.validar(
            indices
        )

        if errores:
            self.errores.extend(errores)
            return self._resultado()

        modalidad, _ = ModalidadAcademica.unfiltered.get_or_create(

            institucion=self.institucion,

            nombre="PRESENCIAL",

        )

        codigo_periodo = "202610"

        if indices["periodo"] is not None:

            fila = next(

                hoja.iter_rows(

                    min_row=2,

                    max_row=2,

                    values_only=True,

                )

            )

            codigo_periodo = str(

                fila[indices["periodo"]]

            ).strip()

        periodo, _ = PeriodoAcademico.unfiltered.get_or_create(

            institucion=self.institucion,

            codigo_institucional=codigo_periodo,

            defaults={

                "modalidad": modalidad,

                "nombre": codigo_periodo,

                "fecha_inicio": datetime(
                    2026,
                    1,
                    1
                ).date(),

                "fecha_fin": datetime(
                    2026,
                    12,
                    31
                ).date(),

            }

                )

        momento, _ = MomentoAcademico.objects.get_or_create(

            periodo=periodo,

            nombre="Periodo Completo",

            defaults={

                "fecha_inicio": periodo.fecha_inicio,

                "fecha_fin": periodo.fecha_fin,

            }

        )


        logger.info("Validando e importando clases...")

        dias = self.DIAS

        clases_por_crear = []

        asignador = AsignadorAulas()

        for numero_fila, fila in enumerate(hoja.iter_rows(
            min_row=2,
            values_only=True,
        ), start=2):

            if not fila or not any(valor not in (None, "") for valor in fila):
                continue

            self.filas_revisadas += 1

            nrc = self._texto(fila[indices["nrc"]])
            asignatura = self._texto(fila[indices["asignatura"]])
            estado_nrc = self._texto(
                fila[indices["estado_nrc"]]
                if indices.get("estado_nrc") is not None else ""
            ).upper()
            nombre_edificio_original = self._texto(fila[indices["edificio"]]).upper()
            nombre_aula_original = self._texto(fila[indices["aula"]]).upper()
            valor_hora_inicio = fila[indices["hora_inicio"]]
            valor_hora_fin = fila[indices["hora_fin"]]
            dias_marcados = [
                nombre_dia
                for nombre_dia in dias
                if indices[nombre_dia] is not None
                and self._dia_marcado(fila[indices[nombre_dia]])
            ]

            # El archivo institucional puede incluir filas de consulta,
            # clases inactivas o actividades aún sin día asignado. No son
            # errores de formato ni crean sedes, aulas o docentes vacíos.
            if nrc.upper().startswith("CONSULTAR"):
                self._omitir_fila(numero_fila, nrc, "referencia")
                continue
            if estado_nrc in {"INACTIVO", "CANCELADO", "CANCELADA"}:
                self._omitir_fila(numero_fila, nrc, "inactiva")
                continue

            tiene_inicio = self._tiene_valor(valor_hora_inicio)
            tiene_fin = self._tiene_valor(valor_hora_fin)
            tiene_horas = tiene_inicio and tiene_fin
            es_virtual = nombre_edificio_original in self.EDIFICIOS_VIRTUALES
            tiene_ubicacion = es_virtual or (
                bool(nombre_edificio_original) and bool(nombre_aula_original)
            )

            if not tiene_inicio and not tiene_fin and not dias_marcados:
                self._omitir_fila(numero_fila, nrc, "sin_programacion")
                continue
            if not dias_marcados:
                self._omitir_fila(numero_fila, nrc, "sin_dia")
                continue
            if not tiene_horas:
                self._omitir_fila(numero_fila, nrc, "sin_horas")
                continue
            if not tiene_ubicacion:
                self._omitir_fila(numero_fila, nrc, "sin_ubicacion")
                continue

            nombre_sede = str(
                fila[indices["sede"]] or ""
            ).strip().upper()

            nombre_sede = self.MAPA_SEDES.get(
                nombre_sede,
                nombre_sede
            )

            sede = Sede.unfiltered.filter(
                institucion=self.institucion,
                nombre__iexact=nombre_sede
            ).first()

            if not sede:

                if not nombre_sede:
                    self.errores.append(
                        f"Fila {numero_fila}: la sede está vacía."
                    )
                    self.filas_invalidas += 1
                    continue

                sede = Sede.unfiltered.create(
                    institucion=self.institucion,
                    nombre=nombre_sede,
                )

                self.sedes_creadas.append(nombre_sede)
                logger.info(f"Sede creada: {nombre_sede}")

            # -----------------------------
            # EDIFICIO
            # -----------------------------

            nombre_edificio = str(
                fila[indices["edificio"]] or ""
            ).strip().upper()

            nombre_edificio = self.MAPA_EDIFICIOS.get(
                nombre_edificio,
                nombre_edificio
            )

            if nombre_edificio in self.EDIFICIOS_VIRTUALES:

                edificio, _ = Edificio.unfiltered.get_or_create(
                    institucion=self.institucion,
                    sede=sede,
                    nombre="Virtual",
                )

                aula, _ = Aula.unfiltered.get_or_create(
                    institucion=self.institucion,
                    edificio=edificio,
                    nombre="VIRTUAL",
                    defaults={
                        "capacidad": 999,
                        "tipo_espacio": "VIRTUAL",
                    }
                )

            else:

                edificio = Edificio.unfiltered.filter(
                    institucion=self.institucion,
                    sede=sede,
                    nombre__iexact=nombre_edificio
                ).first()

                if not edificio:

                    if not nombre_edificio:
                        self.errores.append(
                            f"Fila {numero_fila}: el edificio está vacío "
                            f"(sede: {sede.nombre})."
                        )
                        self.filas_invalidas += 1
                        continue

                    edificio = Edificio.unfiltered.create(
                        institucion=self.institucion,
                        sede=sede,
                        nombre=nombre_edificio,
                    )

                    self.edificios_creados.append(
                        f"{nombre_edificio} ({sede.nombre})"
                    )

                # -----------------------------
                # AULA
                # -----------------------------

                nombre_aula = str(
                    fila[indices["aula"]] or ""
                ).strip().upper()

                nombre_aula = self.MAPA_AULAS.get(
                    nombre_aula,
                    nombre_aula
                )

                aula = Aula.unfiltered.filter(
                    institucion=self.institucion,
                    edificio=edificio,
                    nombre__iexact=nombre_aula
                ).first()

                if not aula:

                    aula = Aula.unfiltered.filter(
                        institucion=self.institucion,
                        nombre__iexact=nombre_aula
                    ).first()

                if not aula:
                    if not nombre_aula:
                        self.errores.append(
                            f"Fila {numero_fila}: el espacio está vacío "
                            f"(edificio: {edificio.nombre})."
                        )
                        self.filas_invalidas += 1
                        continue

                    # Una instituci\u00f3n nueva puede iniciar desde su programaci\u00f3n.
                    # Los datos del espacio se completan despu\u00e9s desde administraci\u00f3n.
                    aula = Aula.unfiltered.create(
                        institucion=self.institucion,
                        edificio=edificio,
                        nombre=nombre_aula,
                        capacidad=30,
                        tipo_espacio="AULA",
                    )
                    self.aulas_creadas.append(
                        f"{nombre_aula} ({edificio.nombre})"
                    )

            # -----------------------------
            # DOCENTE
            # -----------------------------

            nombre_docente = str(
                fila[indices["docente"]] or "SIN DOCENTE"
            ).strip()

            docente, docente_creado = Docente.unfiltered.get_or_create(
                institucion=self.institucion,
                nombre=nombre_docente,
                defaults={
                    "identificacion": "",
                    "email": "",
                }
            )
            if docente_creado:
                self.docentes_creados.append(nombre_docente)

            try:
                hora_inicio = self.convertir_hora(valor_hora_inicio)
                hora_fin = self.convertir_hora(valor_hora_fin)
            except (TypeError, ValueError):
                mensaje = (
                    f"Fila {numero_fila}: las horas no son válidas. "
                    "Usa HHMM, por ejemplo 0830."
                )
                logger.warning(mensaje)
                self.errores.append(mensaje)
                self.filas_invalidas += 1
                continue

            if not hora_inicio or not hora_fin or hora_inicio >= hora_fin:
                mensaje = (
                    f"Fila {numero_fila}: el rango de horas no es válido "
                    f"({hora_inicio}-{hora_fin})."
                )
                logger.warning(mensaje)
                self.errores.append(mensaje)
                self.filas_invalidas += 1
                continue

            if not nrc or not asignatura:
                self.errores.append(
                    f"Fila {numero_fila}: el identificador de clase y la asignatura son obligatorios."
                )
                self.filas_invalidas += 1
                continue

            clases_antes_de_la_fila = len(clases_por_crear)

            for nombre_dia, numero_dia in dias.items():

                indice = indices[nombre_dia]

                if indice is None:
                    continue

                if not self._dia_marcado(fila[indice]):
                    continue

                try:

                    clase = Clase(
                        institucion=self.institucion,
                        docente=docente,
                        aula=aula,
                        periodo=periodo,
                        momento=momento,
                        asignatura=asignatura,
                        nrc=nrc,
                        dia_semana=numero_dia,
                        hora_inicio=hora_inicio,
                        hora_fin=hora_fin,
                    )
                    clases_por_crear.append(clase)

                    # Log first few classes for debugging
                    if len(clases_por_crear) <= 5:
                        logger.info(f"Clase a crear: {asignatura} - {nombre_docente} - {aula.nombre} - {nombre_dia} {hora_inicio}-{hora_fin}")

                except (ValidationError, Exception) as e:

                    self.errores.append(
                        f"Fila {numero_fila}, clase {nrc}: {str(e)}"
                    )

            if len(clases_por_crear) > clases_antes_de_la_fila:
                self.filas_validas += 1
            else:
                self.filas_invalidas += 1

        if self.errores:
            logger.warning("La importación se canceló por errores de validación: %s", len(self.errores))
            transaction.set_rollback(True)
            return self._resultado(total=len(clases_por_crear))

        if not clases_por_crear:
            logger.warning("No se encontraron clases válidas para importar")
            transaction.set_rollback(True)
            self.errores.append("El archivo no contiene clases válidas para importar.")
            return self._resultado()

        # Solo se reemplaza la programación vigente una vez el archivo pasó
        # todas las validaciones necesarias.
        logger.info(f"Preparando para crear {len(clases_por_crear)} clases")
        respaldo_anterior = capturar_programacion(self.institucion)
        Clase.unfiltered.filter(institucion=self.institucion).delete()
        Clase.unfiltered.bulk_create(clases_por_crear)
        self.total = len(clases_por_crear)

        ImportacionProgramacion.unfiltered.create(
            institucion=self.institucion,
            archivo_nombre=getattr(self.archivo, 'name', 'programacion.xlsx'),
            creado_por=self.usuario if getattr(self.usuario, 'is_authenticated', False) else None,
            total_clases=self.total,
            respaldo_anterior=respaldo_anterior,
        )

        logger.info(f"Importación exitosa:")
        logger.info(f"  - Clases creadas: {self.total}")
        logger.info(f"  - Sedes creadas: {len(self.sedes_creadas)}")
        logger.info(f"  - Edificios creados: {len(self.edificios_creados)}")
        logger.info(f"  - Aulas creadas: {len(self.aulas_creadas)}")
        logger.info(f"  - Filas omitidas: {self.filas_omitidas}")

        return self._resultado()
