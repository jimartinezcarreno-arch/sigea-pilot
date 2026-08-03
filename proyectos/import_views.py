from io import BytesIO
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

from django.contrib import messages
from django.core.files.base import ContentFile, File
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from .models import Clase, ImportacionProgramacion, Institucion
from .permissions import roles_requeridos
from .services.excel_importer import ExcelImporter
from .services.programacion_backup import capturar_programacion, clases_desde_respaldo
from .tenant_utils import get_current_institucion


IMPORTACION_PENDIENTE_SESSION_KEY = 'importacion_programacion_pendiente'
TAMANO_MAXIMO_IMPORTACION = 10 * 1024 * 1024


def _institucion_activa():
    return get_current_institucion() or Institucion.objects.first()


def _eliminar_importacion_pendiente(request):
    pendiente = request.session.pop(IMPORTACION_PENDIENTE_SESSION_KEY, None)
    if pendiente and pendiente.get('ruta') and default_storage.exists(pendiente['ruta']):
        default_storage.delete(pendiente['ruta'])
    request.session.modified = True


def _guardar_importacion_pendiente(request, archivo, institucion):
    _eliminar_importacion_pendiente(request)
    extension = Path(archivo.name).suffix.lower()
    if extension != '.xlsx':
        raise ValueError('SIGEA acepta archivos Excel con extensión .xlsx.')
    if archivo.size > TAMANO_MAXIMO_IMPORTACION:
        raise ValueError('El archivo supera el límite de 10 MB permitido para una importación.')

    ruta = default_storage.save(
        f'importaciones-pendientes/{uuid4().hex}{extension}',
        ContentFile(archivo.read()),
    )
    pendiente = {
        'ruta': ruta,
        'nombre': Path(archivo.name).name,
        'institucion_id': institucion.id,
    }
    request.session[IMPORTACION_PENDIENTE_SESSION_KEY] = pendiente
    request.session.modified = True
    return pendiente


def _obtener_importacion_pendiente(request, institucion):
    pendiente = request.session.get(IMPORTACION_PENDIENTE_SESSION_KEY)
    if not pendiente or pendiente.get('institucion_id') != institucion.id:
        return None
    if not pendiente.get('ruta') or not default_storage.exists(pendiente['ruta']):
        _eliminar_importacion_pendiente(request)
        return None
    return pendiente


def _contexto_revision_importacion(institucion, pendiente, resultado):
    return {
        'institucion_activa': institucion,
        'archivo_nombre': pendiente['nombre'],
        'resultado': resultado,
        'clases_actuales': Clase.unfiltered.filter(institucion=institucion).count(),
        'puede_confirmar': not resultado.get('errores') and resultado.get('total', 0) > 0,
    }


@roles_requeridos('ADMIN', 'PROGRAMADOR', 'CONSULTA')
def descargar_plantilla_programacion(request):
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Programación'
    encabezados = [
        'PERIODO', 'ID_CLASE', 'ASIGNATURA', 'DOCENTE', 'SEDE', 'EDIFICIO',
        'ESPACIO', 'HORA_INICIO', 'HORA_FIN', 'LUN', 'MAR', 'MIE', 'JUE',
        'VIE', 'SAB', 'DOM',
    ]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='2563EB')
    hoja.freeze_panes = 'A2'
    hoja.auto_filter.ref = f'A1:{hoja.cell(row=1, column=len(encabezados)).coordinate}'
    hoja.sheet_view.showGridLines = False
    for columna in hoja.columns:
        hoja.column_dimensions[columna[0].column_letter].width = 20

    instrucciones = libro.create_sheet('Instrucciones')
    instrucciones.append(['Plantilla estándar de programación SIGEA'])
    instrucciones['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    instrucciones['A1'].fill = PatternFill('solid', fgColor='1D4ED8')
    instrucciones.merge_cells('A1:B1')
    instrucciones.append(['Campo', 'Uso'])
    instrucciones.append(['PERIODO', 'Código o nombre del periodo académico, por ejemplo 2026-2.'])
    instrucciones.append(['ID_CLASE', 'Identificador único de la clase. Puede ser NRC, consecutivo o código institucional.'])
    instrucciones.append(['ASIGNATURA', 'Nombre de la asignatura, curso o actividad académica.'])
    instrucciones.append(['DOCENTE', 'Nombre completo de la persona responsable de la clase.'])
    instrucciones.append(['SEDE / EDIFICIO / ESPACIO', 'Ubicación de la clase. SIGEA crea el catálogo faltante con capacidad provisional de 30 personas.'])
    instrucciones.append(['HORA_INICIO / HORA_FIN', 'Hora de inicio y fin en formato HHMM: por ejemplo 0830 y 1000.'])
    instrucciones.append(['LUN a DOM', 'Marca con X los días en que ocurre la clase.'])
    instrucciones.append(['Importante', 'Completa solo la hoja Programación. Las hojas Instrucciones y Ejemplo no se importan.'])
    instrucciones.append(['Compatibilidad', 'SIGEA también reconoce encabezados habituales como NRC, TITULO, NOMBRE_DOCENTE, SALON, HI, HF y L/M/I/J/V/S/D.'])
    instrucciones.column_dimensions['A'].width = 34
    instrucciones.column_dimensions['B'].width = 100
    for celda in instrucciones[2]:
        celda.font = Font(bold=True)

    ejemplo = libro.create_sheet('Ejemplo')
    ejemplo.append(encabezados)
    ejemplo.append([
        '2026-2', 'MAT-101-01', 'Matemáticas I', 'Ana Pérez', 'Sede principal',
        'Edificio A', 'A-101', '0800', '1000', 'X', '', '', '', '', '', '',
    ])
    for celda in ejemplo[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='0F766E')
    ejemplo.freeze_panes = 'A2'
    ejemplo.sheet_view.showGridLines = False
    for columna in ejemplo.columns:
        ejemplo.column_dimensions[columna[0].column_letter].width = 20

    contenido = BytesIO()
    libro.save(contenido)
    respuesta = HttpResponse(
        contenido.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    respuesta['Content-Disposition'] = 'attachment; filename="plantilla_programacion_sigea.xlsx"'
    return respuesta


@roles_requeridos('ADMIN', 'PROGRAMADOR')
@require_POST
def subir_programacion_segura(request):
    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Selecciona un archivo Excel antes de importar.')
        return redirect('consulta_aulas')

    institucion = _institucion_activa()
    if institucion is None:
        messages.error(request, 'No hay una institución activa para recibir la programación.')
        return redirect('consulta_aulas')

    try:
        pendiente = _guardar_importacion_pendiente(request, archivo, institucion)
        with default_storage.open(pendiente['ruta'], 'rb') as archivo_pendiente:
            archivo_revision = File(archivo_pendiente, name=pendiente['nombre'])
            resultado = ExcelImporter(
                archivo_revision, institucion, usuario=request.user
            ).previsualizar()
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as error:
        _eliminar_importacion_pendiente(request)
        messages.error(request, f'No fue posible revisar el archivo: {error}')
        return redirect('consulta_aulas')

    contexto = _contexto_revision_importacion(institucion, pendiente, resultado)
    if resultado.get('errores'):
        _eliminar_importacion_pendiente(request)
        contexto['puede_confirmar'] = False
    return render(request, 'confirmar_importacion.html', contexto)


@roles_requeridos('ADMIN', 'PROGRAMADOR')
@require_POST
def confirmar_importacion(request):
    institucion = _institucion_activa()
    pendiente = _obtener_importacion_pendiente(request, institucion) if institucion else None
    if pendiente is None:
        messages.error(request, 'No hay una importación pendiente para confirmar. Revisa el archivo nuevamente.')
        return redirect('consulta_aulas')

    try:
        with default_storage.open(pendiente['ruta'], 'rb') as archivo_pendiente:
            archivo_importacion = File(archivo_pendiente, name=pendiente['nombre'])
            resultado = ExcelImporter(
                archivo_importacion, institucion, usuario=request.user
            ).importar()
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as error:
        messages.error(request, f'No fue posible aplicar la importación: {error}')
        return redirect('consulta_aulas')
    finally:
        _eliminar_importacion_pendiente(request)

    if resultado.get('errores'):
        for error in resultado['errores'][:20]:
            messages.warning(request, error)
        messages.error(request, 'La programación vigente se conservó; corrige el archivo e intenta de nuevo.')
        return redirect('consulta_aulas')

    for clave, etiqueta in (
        ('sedes_creadas', 'Sedes creadas'),
        ('edificios_creados', 'Edificios creados'),
        ('aulas_creadas', 'Espacios creados con capacidad provisional de 30 personas'),
        ('docentes_creados', 'Docentes creados'),
    ):
        if resultado.get(clave):
            messages.info(request, f"{etiqueta}: " + ', '.join(sorted(set(resultado[clave]))))

    messages.success(
        request,
        f"Importación completada: {resultado['total']} clases. El estado anterior quedó guardado en el historial.",
    )
    return redirect('consulta_aulas')


@roles_requeridos('ADMIN', 'PROGRAMADOR')
@require_POST
def cancelar_importacion(request):
    _eliminar_importacion_pendiente(request)
    messages.info(request, 'La importación pendiente fue descartada. La programación vigente no cambió.')
    return redirect('consulta_aulas')


@roles_requeridos('ADMIN')
def historial_importaciones(request):
    institucion = _institucion_activa()
    if institucion is None:
        raise Http404('No hay institución activa.')
    importaciones = ImportacionProgramacion.objects.select_related('creado_por').all()
    return render(request, 'historial_importaciones.html', {
        'institucion_activa': institucion,
        'importaciones': importaciones,
    })


@roles_requeridos('ADMIN')
@require_POST
def restaurar_importacion(request, importacion_id):
    try:
        importacion = ImportacionProgramacion.objects.get(id=importacion_id)
    except ImportacionProgramacion.DoesNotExist as error:
        raise Http404('No se encontró la importación solicitada.') from error

    institucion = _institucion_activa()
    respaldo = importacion.respaldo_anterior
    try:
        clases_restauradas = clases_desde_respaldo(institucion, respaldo)
    except (KeyError, TypeError, ValueError):
        messages.error(request, 'El respaldo no tiene un formato válido y no se modificó la programación.')
        return redirect('historial_importaciones')

    with transaction.atomic():
        respaldo_actual = capturar_programacion(institucion)
        Clase.unfiltered.filter(institucion=institucion).delete()
        Clase.unfiltered.bulk_create(clases_restauradas)
        ImportacionProgramacion.unfiltered.create(
            institucion=institucion,
            archivo_nombre=f'Restauración de #{importacion.id}: {importacion.archivo_nombre}',
            tipo='RESTAURACION',
            creado_por=request.user,
            total_clases=len(clases_restauradas),
            respaldo_anterior=respaldo_actual,
        )

    messages.success(request, f'Se restauraron {len(clases_restauradas)} clases de la versión seleccionada.')
    return redirect('historial_importaciones')
