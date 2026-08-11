import os
from datetime import date, time
from io import BytesIO
from unittest.mock import patch

from django.test import TestCase, override_settings
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .models import Aula, Clase, Docente, Edificio, ImportacionProgramacion, Institucion, ModalidadAcademica, MomentoAcademico, PerfilUsuario, PeriodoAcademico, Sede
from .services.excel_importer import ExcelImporter
from .services.programacion_backup import capturar_programacion
from .tenant_utils import set_current_institucion


@override_settings(ALLOWED_HOSTS=["testserver", "inst1.localhost", "unknown.localhost"])
class SIGEATestCase(TestCase):
    def setUp(self):
        self.inst1 = Institucion.objects.create(
            nombre="Institucion Uno", subdominio="inst1",
            hora_inicio_jornada=time(6, 0), hora_fin_jornada=time(22, 0),
        )
        self.inst2 = Institucion.objects.create(
            nombre="Institucion Dos", subdominio="inst2",
            hora_inicio_jornada=time(6, 0), hora_fin_jornada=time(22, 0),
        )
        set_current_institucion(self.inst1)
        modalidad = ModalidadAcademica.objects.create(institucion=self.inst1, nombre="Pregrado")
        periodo = PeriodoAcademico.objects.create(
            institucion=self.inst1, modalidad=modalidad, codigo_institucional="2026-1", nombre="2026-1",
            fecha_inicio=date(2026, 1, 1), fecha_fin=date(2026, 12, 31),
        )
        self.momento = MomentoAcademico.objects.create(
            periodo=periodo, nombre="Periodo completo",
            fecha_inicio=date(2026, 1, 1), fecha_fin=date(2026, 12, 31),
        )
        sede = Sede.objects.create(institucion=self.inst1, nombre="Sede principal")
        edificio = Edificio.objects.create(institucion=self.inst1, sede=sede, nombre="Edificio A")
        self.aula = Aula.objects.create(institucion=self.inst1, edificio=edificio, nombre="A-101", capacidad=30, tipo_espacio="AULA")
        self.aula_alterna = Aula.objects.create(institucion=self.inst1, edificio=edificio, nombre="A-102", capacidad=30, tipo_espacio="AULA")
        self.docente = Docente.objects.create(institucion=self.inst1, nombre="Ana Perez", identificacion="1", email="ana@example.com")
        self.clase = Clase.objects.create(
            institucion=self.inst1, docente=self.docente, aula=self.aula, periodo=periodo, momento=self.momento,
            asignatura="Matematicas", nrc="MAT-1", dia_semana=1,
            hora_inicio=time(8, 0), hora_fin=time(10, 0),
        )
        set_current_institucion(None)

    @staticmethod
    def archivo_programacion(nrc='MAT-2', hora_inicio='0800'):
        libro = Workbook()
        hoja = libro.active
        hoja.append([
            'PERIODO', 'ID_CLASE', 'ASIGNATURA', 'DOCENTE', 'SEDE', 'EDIFICIO',
            'ESPACIO', 'HORA_INICIO', 'HORA_FIN', 'LUN', 'MAR', 'MIE', 'JUE',
            'VIE', 'SAB', 'DOM',
        ])
        hoja.append([
            '2026-2', nrc, 'Álgebra', 'Ana Pérez', 'Sede principal', 'Edificio A',
            'A-101', hora_inicio, '1000', 'X', '', '', '', '', '', '',
        ])
        contenido = BytesIO()
        libro.save(contenido)
        return SimpleUploadedFile(
            'programacion-prueba.xlsx', contenido.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_unknown_tenant_cannot_read_tenant_data(self):
        self.assertEqual(Aula.objects.count(), 0)
        respuesta = self.client.get("/", HTTP_HOST="unknown.localhost")
        self.assertRedirects(respuesta, "/acceso/?next=/")

    def test_agenda_contains_recurring_calendar_event(self):
        respuesta = self.client.get(f"/agenda-docente/{self.docente.id}/", HTTP_HOST="inst1.localhost")
        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, '"daysOfWeek": [1]')
        self.assertContains(respuesta, "Matematicas")

    def test_aula_sin_programacion_se_diferencia_sin_dejar_de_estar_disponible(self):
        respuesta = self.client.get('/consultar-aulas/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 200)
        tarjetas = {item['aula'].nombre: item for item in respuesta.context['resultado']}

        self.assertFalse(tarjetas['A-101']['sin_programacion'])
        self.assertTrue(tarjetas['A-102']['sin_programacion'])
        self.assertEqual(tarjetas['A-102']['estado'], 'DISPONIBLE')
        self.assertEqual(tarjetas['A-101']['programacion_total'], 1)
        self.assertEqual(tarjetas['A-102']['programacion_total'], 0)
        self.assertContains(respuesta, 'Sin programación registrada')
        self.assertContains(respuesta, 'Ver programación completa')

    def test_reassignment_requires_post_and_uses_available_room(self):
        url = f"/conflictos/reasignar/{self.clase.id}/{self.aula_alterna.id}/"
        respuesta = self.client.get(url, HTTP_HOST="inst1.localhost")
        self.assertEqual(respuesta.status_code, 405)
        respuesta = self.client.post(url, HTTP_HOST="inst1.localhost")
        self.assertEqual(respuesta.status_code, 302)
        self.clase.refresh_from_db()
        self.assertEqual(self.clase.aula_id, self.aula_alterna.id)

    @override_settings(REQUIRE_LOGIN=True)
    def test_pilot_access_requires_an_account_from_the_current_institution(self):
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertRedirects(respuesta, '/acceso/?next=/')

        user = get_user_model().objects.create_user(username='consulta', password='clave-segura')
        PerfilUsuario.objects.create(user=user, institucion=self.inst2, rol='CONSULTA')
        self.client.force_login(user)
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 403)

        user.perfil_sigea.institucion = self.inst1
        user.perfil_sigea.save(update_fields=['institucion'])
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 200)

    @override_settings(REQUIRE_LOGIN=True)
    def test_roles_protegen_carga_y_gestion_de_usuarios(self):
        consulta = get_user_model().objects.create_user(username='consulta-rol', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=consulta, institucion=self.inst1, rol='CONSULTA')
        self.client.force_login(consulta)

        respuesta = self.client.get('/usuarios/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 403)
        self.assertContains(
            respuesta,
            'Solo el administrador institucional puede gestionar usuarios.',
            status_code=403,
        )
        respuesta = self.client.post('/subir-excel/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 403)

        programador = get_user_model().objects.create_user(username='programador', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=programador, institucion=self.inst1, rol='PROGRAMADOR')
        self.client.force_login(programador)
        respuesta = self.client.post('/subir-excel/', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 302)

    @override_settings(REQUIRE_LOGIN=True)
    def test_administrador_puede_crear_cuenta_de_su_institucion(self):
        admin = get_user_model().objects.create_user(username='admin-rol', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=admin, institucion=self.inst1, rol='ADMIN')
        self.client.force_login(admin)

        respuesta = self.client.post('/usuarios/', {
            'accion': 'crear',
            'username': 'piloto-consulta',
            'email': 'consulta@example.com',
            'rol': 'CONSULTA',
            'password': 'OtraClaveSegura123!',
        }, HTTP_HOST='inst1.localhost')
        self.assertRedirects(respuesta, '/usuarios/')
        perfil = PerfilUsuario.objects.get(user__username='piloto-consulta')
        self.assertEqual(perfil.institucion, self.inst1)
        self.assertEqual(perfil.rol, 'CONSULTA')

        respuesta = self.client.get('/usuarios/', HTTP_HOST='inst1.localhost')
        self.assertContains(respuesta, 'Programador académico')

    @override_settings(REQUIRE_LOGIN=True)
    def test_programador_revisa_y_confirma_antes_de_reemplazar_programacion(self):
        programador = get_user_model().objects.create_user(
            username='programador-importa', password='ClaveSegura123!'
        )
        PerfilUsuario.objects.create(user=programador, institucion=self.inst1, rol='PROGRAMADOR')
        self.client.force_login(programador)

        respuesta = self.client.post('/subir-excel/', {
            'archivo_excel': self.archivo_programacion(),
        }, HTTP_HOST='inst1.localhost')

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'REVISIÓN DE PROGRAMACIÓN')
        self.assertContains(respuesta, 'Confirmar y reemplazar programación')
        self.assertEqual(Clase.unfiltered.get(institucion=self.inst1).nrc, 'MAT-1')
        self.assertEqual(ImportacionProgramacion.unfiltered.filter(institucion=self.inst1).count(), 0)

        respuesta = self.client.post('/importaciones/confirmar/', HTTP_HOST='inst1.localhost')
        self.assertRedirects(respuesta, '/consultar-aulas/')
        self.assertEqual(Clase.unfiltered.get(institucion=self.inst1).nrc, 'MAT-2')
        self.assertEqual(ImportacionProgramacion.unfiltered.filter(institucion=self.inst1).count(), 1)

    @override_settings(REQUIRE_LOGIN=True)
    def test_revision_con_errores_no_reemplaza_la_programacion(self):
        programador = get_user_model().objects.create_user(
            username='programador-revision', password='ClaveSegura123!'
        )
        PerfilUsuario.objects.create(user=programador, institucion=self.inst1, rol='PROGRAMADOR')
        self.client.force_login(programador)

        respuesta = self.client.post('/subir-excel/', {
            'archivo_excel': self.archivo_programacion(hora_inicio='hora-invalida'),
        }, HTTP_HOST='inst1.localhost')

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'No se realizó ningún cambio')
        self.assertNotContains(respuesta, 'Confirmar y reemplazar programación')
        self.assertEqual(Clase.unfiltered.get(institucion=self.inst1).nrc, 'MAT-1')

    @override_settings(REQUIRE_LOGIN=True)
    def test_cierre_de_sesion_requiere_post_y_redirige_al_acceso(self):
        user = get_user_model().objects.create_user(username='cerrar-sesion', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=user, institucion=self.inst1, rol='CONSULTA')
        self.client.force_login(user)

        respuesta = self.client.post('/salir/', HTTP_HOST='inst1.localhost')
        self.assertRedirects(respuesta, '/acceso/')

    @override_settings(REQUIRE_LOGIN=True)
    def test_navegacion_global_muestra_inicio_y_cierre_de_sesion(self):
        usuario = get_user_model().objects.create_user(
            username='navegacion', password='ClaveSegura123!'
        )
        PerfilUsuario.objects.create(user=usuario, institucion=self.inst1, rol='CONSULTA')
        self.client.force_login(usuario)

        inicio = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertContains(inicio, 'Cerrar sesión')
        self.assertContains(inicio, 'action="/salir/"')

        consulta = self.client.get('/consultar-aulas/', HTTP_HOST='inst1.localhost')
        self.assertContains(consulta, 'Módulos')
        self.assertContains(consulta, f'href="{reverse("dashboard_modulos")}"')
        self.assertContains(consulta, 'Cerrar sesión')

    @override_settings(REQUIRE_LOGIN=True)
    def test_administrador_puede_restaurar_programacion_previa(self):
        respaldo = capturar_programacion(self.inst1)
        importacion = ImportacionProgramacion.unfiltered.create(
            institucion=self.inst1,
            archivo_nombre='programacion-anterior.xlsx',
            total_clases=1,
            respaldo_anterior=respaldo,
        )
        self.clase.nrc = 'VERSION-ACTUAL'
        self.clase.save(update_fields=['nrc'])

        admin = get_user_model().objects.create_user(username='admin-restaura', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=admin, institucion=self.inst1, rol='ADMIN')
        self.client.force_login(admin)

        respuesta = self.client.post(
            f'/importaciones/{importacion.id}/restaurar/',
            HTTP_HOST='inst1.localhost',
        )
        self.assertRedirects(respuesta, '/importaciones/')
        self.assertEqual(Clase.unfiltered.get(institucion=self.inst1).nrc, 'MAT-1')
        self.assertEqual(
            ImportacionProgramacion.unfiltered.filter(institucion=self.inst1, tipo='RESTAURACION').count(),
            1,
        )

    @override_settings(REQUIRE_LOGIN=True)
    def test_programador_puede_descargar_plantilla_de_importacion(self):
        programador = get_user_model().objects.create_user(username='plantilla', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=programador, institucion=self.inst1, rol='PROGRAMADOR')
        self.client.force_login(programador)

        respuesta = self.client.get('/plantilla-programacion.xlsx', HTTP_HOST='inst1.localhost')
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        libro = load_workbook(BytesIO(respuesta.content))
        self.assertEqual(libro['Programación']['A1'].value, 'PERIODO')
        self.assertEqual(libro['Programación']['B1'].value, 'ID_CLASE')
        self.assertEqual(libro['Programación']['P1'].value, 'DOM')
        self.assertIn('Instrucciones', libro.sheetnames)
        self.assertIn('Ejemplo', libro.sheetnames)

    @override_settings(REQUIRE_LOGIN=True)
    def test_tablero_muestra_herramientas_segun_el_rol(self):
        consulta = get_user_model().objects.create_user(username='vista-consulta', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=consulta, institucion=self.inst1, rol='CONSULTA')
        self.client.force_login(consulta)
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertContains(respuesta, 'Modo consulta')
        self.assertNotContains(respuesta, 'Importar Excel')
        self.assertNotContains(respuesta, 'Gestión de Usuarios')

        programador = get_user_model().objects.create_user(username='vista-programador', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=programador, institucion=self.inst1, rol='PROGRAMADOR')
        self.client.force_login(programador)
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertContains(respuesta, 'Importar Excel')
        self.assertNotContains(respuesta, 'Gestión de Usuarios')

        admin = get_user_model().objects.create_user(username='vista-admin', password='ClaveSegura123!')
        PerfilUsuario.objects.create(user=admin, institucion=self.inst1, rol='ADMIN')
        self.client.force_login(admin)
        respuesta = self.client.get('/', HTTP_HOST='inst1.localhost')
        self.assertContains(respuesta, 'Gestión de Usuarios')


class BootstrapPilotCommandTests(TestCase):
    @patch.dict(os.environ, {
        'DEFAULT_TENANT_SUBDOMAIN': 'piloto-prueba',
        'INITIAL_ADMIN_USERNAME': 'admin-piloto',
        'INITIAL_ADMIN_PASSWORD': 'ClaveSegura123!',
        'INITIAL_ADMIN_EMAIL': 'admin@example.com',
    }, clear=False)
    def test_crea_institucion_y_administrador_de_forma_idempotente(self):
        call_command('bootstrap_pilot')
        call_command('bootstrap_pilot')

        institucion = Institucion.objects.get(subdominio='piloto-prueba')
        usuario = get_user_model().objects.get(username='admin-piloto')
        self.assertTrue(usuario.is_superuser)
        self.assertTrue(usuario.check_password('ClaveSegura123!'))
        self.assertEqual(usuario.perfil_sigea.institucion, institucion)
        self.assertEqual(usuario.perfil_sigea.rol, 'ADMIN')


class ExcelImporterTests(TestCase):
    def setUp(self):
        self.institucion = Institucion.objects.create(
            nombre='Institucion de prueba', subdominio='importacion',
            hora_inicio_jornada=time(6, 0), hora_fin_jornada=time(22, 0),
        )

    def crear_archivo(self, hora_inicio='0800', nrc='NRC-1', incluir_fila_invalida=False):
        libro = Workbook()
        hoja = libro.active
        hoja.append([
            'PERIODO', 'NRC', 'TITULO', 'NOMBRE_DOCENTE', 'SEDE', 'EDIFICIO',
            'SALON', 'HI', 'HF', 'L', 'M', 'I', 'J', 'V', 'S', 'D',
        ])
        hoja.append([
            '202610', nrc, 'Matematicas', 'Ana Perez', 'Sede piloto',
            'Edificio A', 'A-101', hora_inicio, '1000', 'X', '', '', '', '', '', '',
        ])
        if incluir_fila_invalida:
            hoja.append([
                '202610', 'NRC-INVALIDO', 'Física', 'Ana Perez', 'Sede piloto',
                'Edificio A', 'A-102', 'hora-invalida', '1000', 'X', '', '', '', '', '', '',
            ])
        from io import BytesIO
        contenido = BytesIO()
        libro.save(contenido)
        return SimpleUploadedFile(
            'programacion.xlsx',
            contenido.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @staticmethod
    def archivo_con_registros_no_programados():
        """Representa el formato institucional: clases, referencias e inactivos."""
        libro = Workbook()
        hoja = libro.active
        hoja.append([
            'PERIODO', 'NRC', 'ESTADO_NRC', 'TITULO', 'NOMBRE_DOCENTE',
            'SEDE', 'EDIFICIO', 'SALON', 'HI', 'HF', 'L', 'M', 'I', 'J',
            'V', 'S', 'D',
        ])
        hoja.append([
            '202645', 'NRC-REAL', 'Activo', 'Matematicas', 'Ana Perez',
            'BUC', 'DJCB', '305', '0830', '1000', 'X', '', '', '', '', '', '',
        ])
        hoja.append([
            '202645', 'Consultar Nacionales', '', 'Ingles III', '', 'BUC',
            '', '', '', '', '', '', '', '', '', '', '',
        ])
        hoja.append([
            '202645', 'NRC-INACTIVO', 'Inactivo', 'Curso cerrado', '', 'BUC',
            '', '', '', '', '', '', '', '', '', '', '',
        ])
        hoja.append([
            '202645', 'NRC-VIRTUAL', 'Activo', 'Actividad virtual', 'Ana Perez',
            'BUC', 'VIRTU', '', '0600', '0759', '', '', '', '', '', '', '',
        ])
        contenido = BytesIO()
        libro.save(contenido)
        return SimpleUploadedFile(
            'planeacion-institucional.xlsx', contenido.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_crea_espacios_faltantes_y_conserva_horario_si_hay_errores(self):
        resultado = ExcelImporter(self.crear_archivo(), self.institucion).importar()
        self.assertEqual(resultado['errores'], [])
        self.assertEqual(resultado['total'], 1)
        aula = Aula.unfiltered.get(institucion=self.institucion, nombre='A-101')
        self.assertEqual(aula.capacidad, 30)
        primera_importacion = ImportacionProgramacion.unfiltered.get(institucion=self.institucion)
        self.assertEqual(primera_importacion.respaldo_anterior, [])

        resultado_reemplazo = ExcelImporter(
            self.crear_archivo(nrc='NRC-2'), self.institucion
        ).importar()
        self.assertEqual(resultado_reemplazo['total'], 1)
        ultima_importacion = ImportacionProgramacion.unfiltered.latest('fecha_creacion')
        self.assertEqual(ultima_importacion.respaldo_anterior[0]['nrc'], 'NRC-1')

        resultado_invalido = ExcelImporter(
            self.crear_archivo(hora_inicio='hora-invalida'), self.institucion
        ).importar()
        self.assertTrue(resultado_invalido['errores'])
        self.assertEqual(Clase.unfiltered.filter(institucion=self.institucion).count(), 1)

    def test_un_archivo_con_una_fila_invalida_no_reemplaza_la_programacion(self):
        ExcelImporter(self.crear_archivo(nrc='NRC-1'), self.institucion).importar()

        resultado = ExcelImporter(
            self.crear_archivo(nrc='NRC-2', incluir_fila_invalida=True), self.institucion
        ).importar()

        self.assertTrue(resultado['errores'])
        self.assertEqual(resultado['filas_validas'], 1)
        self.assertEqual(resultado['filas_invalidas'], 1)
        self.assertEqual(Clase.unfiltered.get(institucion=self.institucion).nrc, 'NRC-1')

    def test_importa_programacion_completa_y_omite_referencias_e_inactivos(self):
        resultado = ExcelImporter(
            self.archivo_con_registros_no_programados(), self.institucion
        ).importar()

        self.assertEqual(resultado['errores'], [])
        self.assertEqual(resultado['filas_revisadas'], 4)
        self.assertEqual(resultado['filas_validas'], 1)
        self.assertEqual(resultado['filas_omitidas'], 3)
        self.assertEqual(resultado['total'], 1)
        self.assertEqual(
            {item['etiqueta']: item['total'] for item in resultado['resumen_omitidas']},
            {
                'Referencias sin programación (Consultar ...)': 1,
                'Clases marcadas como inactivas': 1,
                'Registros sin día de clase': 1,
            },
        )
        self.assertEqual(
            list(Clase.unfiltered.filter(institucion=self.institucion).values_list('nrc', flat=True)),
            ['NRC-REAL'],
        )
        self.assertFalse(Aula.unfiltered.filter(institucion=self.institucion, nombre='VIRTUAL').exists())
